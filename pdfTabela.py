import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Regex Ajustado ────────────────────────────────────────────────────────────
RE_PRODUCT = re.compile(
    r"Produto:\s+(\S+)\s+(.+?)\s+Unidade:\s*(\S+)\s+Embalagem:\s*(\S+)\s*$"
)
RE_SALDO = re.compile(
    r"Saldo\s+Anterior:\s*([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
)

# Alterado: Uso de (\S*) para permitir colunas iniciais vazias e \s* para flexibilidade
RE_TRANSACTION = re.compile(
    r"^(\S*)\s*(\S*)\s*(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\S+)\s+(\S+)"
    r"\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
    r"\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
    r"\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
    r"(?:\s+(.*))?$"
)

RE_TOTAL = re.compile(
    r"N[úu]mero\s+Total\s+de\s+Notas\s+(\d+)"
    r"\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
    r"\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
    r"\s+([\d.,]+)\s+([\d.,]+)"
)

SKIP_STARTS = (
    "REGISTRO DE CONTROLE", "Período de Movimentação", "Controle Permanente",
    "RELATÓRIO A TÍTULO", "Empresa:", "CNPJ:", "Documento", "Lançamento",
    "Esp.", "Série", "Contábil", "Data de Emissão",
)

def br_float(s):
    if not s or s.strip() == "": return 0.0
    try:
        return float(str(s).replace(".", "").replace(",", "."))
    except Exception:
        return s

# ─── Estilos (Mantidos como constantes para performance) ──────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=8, name="Arial")
PRODUCT_FILL = PatternFill("solid", fgColor="D9E1F2")
PRODUCT_FONT = Font(bold=True, size=8, name="Arial")
TOTAL_FILL   = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FONT   = Font(bold=True, size=8, name="Arial")
DATA_FONT    = Font(size=8, name="Arial")
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("Esp.", 6), ("Série e Sub", 10), ("Número", 10),
    ("Dia", 5), ("Mês", 5), ("Ano", 5),
    ("Contábil", 8), ("Fiscal", 8),
    ("Ent.\nQuant.", 10), ("Vl.Unitário", 12), ("Vl.IPI", 8), ("Valor Total", 12),
    ("Saíd.\nQuantid.", 10), ("Vl.Unitário", 12), ("Vl.IPI", 8), ("Valor Total", 12),
    ("Est.\nQuant.", 10), ("Vl.Unit", 12), ("Vl.Total", 12),
    ("Observação", 22),
]

SEC_HEADERS = [
    (1, 3, "Documento"), (4, 8, "Lançamento"),
    (9, 12, "Entradas"), (13, 16, "Saídas"),
    (17, 19, "Estoque"), (20, 20, ""),
]

def apply_style(cell, font=None, fill=None, align=None):
    cell.border = BORDER
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align

# ─── Excel writer ──────────────────────────────────────────────────────────────
class ExcelWriter:
    def __init__(self, path):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Estoque"
        self.path = path
        self.row = 1
        for col, (_, w) in enumerate(COLS, 1):
            self.ws.column_dimensions[get_column_letter(col)].width = w

    def write_product_header(self, prod):
        ws, r = self.ws, self.row
        label = (f"Produto: {prod['code']}  {prod['name']}   "
                 f"Unidade: {prod['unit']}   Embalagem: {prod['embalagem']}   "
                 f"Saldo Anterior: {prod['saldo_q']:,.3f}   "
                 f"{prod['saldo_vu']:,.6f}   {prod['saldo_vt']:,.2f}")
        
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        apply_style(ws.cell(r, 1, label), font=PRODUCT_FONT, fill=PRODUCT_FILL, align=Alignment(horizontal="left"))
        self.row += 1

        # Cabeçalhos de Seção
        r2 = self.row
        for s, e, title in SEC_HEADERS:
            if s != e:
                ws.merge_cells(start_row=r2, start_column=s, end_row=r2, end_column=e)
            apply_style(ws.cell(r2, s, title), font=HEADER_FONT, fill=HEADER_FILL, align=Alignment(horizontal="center"))
        self.row += 1

        # Cabeçalhos de Coluna
        r3 = self.row
        for col, (name, _) in enumerate(COLS, 1):
            apply_style(ws.cell(r3, col, name), font=HEADER_FONT, fill=HEADER_FILL, align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r3].height = 28
        self.row += 1

    def write_transaction(self, fields):
        for col, val in enumerate(fields, 1):
            # Alinhamento dinâmico: Números à direita, texto à esquerda
            aln_h = "right" if isinstance(val, (float, int)) else "center" if col <= 6 else "left"
            apply_style(self.ws.cell(self.row, col, val), font=DATA_FONT, align=Alignment(horizontal=aln_h))
        self.row += 1

    def write_total(self, tf):
        r = self.row
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        apply_style(self.ws.cell(r, 1, f"Número Total de Notas: {tf[0]}"), font=TOTAL_FONT, fill=TOTAL_FILL)
        
        mapping = {9: tf[1], 11: tf[2], 12: tf[3], 13: tf[4], 15: tf[5], 16: tf[6], 17: tf[7], 19: tf[8]}
        for col, val in mapping.items():
            apply_style(self.ws.cell(r, col, val), font=TOTAL_FONT, fill=TOTAL_FILL, align=Alignment(horizontal="right"))
        self.row += 1

    def save(self):
        self.wb.save(self.path)

# ─── PDF parser ────────────────────────────────────────────────────────────────
def parse_pdf(pdf_path, xlsx_path, progress_cb=None):
    writer = ExcelWriter(xlsx_path)
    pending_product = None
    in_tx = False

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        for page_num, page in enumerate(pdf.pages, 1):
            if progress_cb: progress_cb(page_num, total_pages)

            # Aumentar um pouco a tolerância ajuda a manter colunas vazias como espaços
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if not text: continue

            for line in text.splitlines():
                clean_line = line.strip()
                if not clean_line: continue

                # Identificação do Produto
                m_prod = RE_PRODUCT.match(clean_line)
                if m_prod:
                    pending_product = {
                        "code": m_prod.group(1), "name": m_prod.group(2).strip(),
                        "unit": m_prod.group(3), "embalagem": m_prod.group(4),
                        "saldo_q": 0, "saldo_vu": 0, "saldo_vt": 0
                    }
                    in_tx = False
                    continue

                # Saldo Anterior
                m_saldo = RE_SALDO.match(clean_line)
                if m_saldo and pending_product:
                    pending_product["saldo_q"]  = br_float(m_saldo.group(1))
                    pending_product["saldo_vu"] = br_float(m_saldo.group(2))
                    pending_product["saldo_vt"] = br_float(m_saldo.group(3))
                    writer.write_product_header(pending_product)
                    pending_product = None
                    in_tx = True
                    continue

                if any(clean_line.startswith(k) for k in SKIP_STARTS):
                    continue

                if not in_tx: continue

                # Linha de Totalização do Produto
                m_total = RE_TOTAL.match(clean_line)
                if m_total:
                    writer.write_total([m_total.group(1)] + [br_float(m_total.group(i)) for i in range(2, 10)])
                    in_tx = False
                    continue

                # Transação (AQUI ESTAVA O PROBLEMA)
                # Usamos search em vez de match para ser mais tolerante com espaços no início
                m_tx = RE_TRANSACTION.search(line) 
                if m_tx:
                    g = m_tx.groups()
                    # Converte colunas financeiras (índice 8 em diante) para float
                    fields = [br_float(v) if i >= 8 and i <= 18 else v for i, v in enumerate(g)]
                    writer.write_transaction(fields)

    writer.save()

# ─── GUI ───────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Conversor PDF → Excel  |  Estoque MODELO-03")
        self.resizable(False, False)
        self.configure(bg="#F0F4FA")
        self._build()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    def _build(self):
        bar = tk.Frame(self, bg="#2A55A0", pady=14)
        bar.pack(fill="x")
        tk.Label(bar, text="📄  PDF → Excel  |  Controle de Estoque",
                 bg="#2A55A0", fg="white", font=("Arial", 13, "bold")).pack()

        f1 = tk.LabelFrame(self, text=" Arquivo PDF ", bg="#F0F4FA",
                           font=("Arial", 9, "bold"), fg="#2A55A0", bd=1, relief="groove")
        f1.pack(fill="x", padx=18, pady=(14, 4))
        self.pdf_var = tk.StringVar()
        tk.Entry(f1, textvariable=self.pdf_var, width=52, font=("Arial", 9),
                 state="readonly", readonlybackground="white").pack(side="left", padx=(10,4), pady=8)
        tk.Button(f1, text="Selecionar…", command=self._pick_pdf,
                  bg="#2A55A0", fg="white", font=("Arial", 9, "bold"),
                  relief="flat", padx=10, cursor="hand2").pack(side="left", pady=8)

        f2 = tk.LabelFrame(self, text=" Salvar Excel como ", bg="#F0F4FA",
                           font=("Arial", 9, "bold"), fg="#2A55A0", bd=1, relief="groove")
        f2.pack(fill="x", padx=18, pady=4)
        self.xlsx_var = tk.StringVar()
        tk.Entry(f2, textvariable=self.xlsx_var, width=52, font=("Arial", 9),
                 state="readonly", readonlybackground="white").pack(side="left", padx=(10,4), pady=8)
        tk.Button(f2, text="Alterar…", command=self._pick_xlsx,
                  bg="#5B7FBF", fg="white", font=("Arial", 9, "bold"),
                  relief="flat", padx=10, cursor="hand2").pack(side="left", pady=8)

        f3 = tk.Frame(self, bg="#F0F4FA")
        f3.pack(fill="x", padx=18, pady=(6, 0))
        self.progress = ttk.Progressbar(f3, mode="determinate", length=460, maximum=100)
        self.progress.pack(fill="x")
        self.status_var = tk.StringVar(value="Selecione um arquivo PDF para começar.")
        tk.Label(f3, textvariable=self.status_var, bg="#F0F4FA",
                 font=("Arial", 8), fg="#555").pack(anchor="w", pady=(3, 0))

        self.btn = tk.Button(self, text="⚡  Converter", command=self._start,
                             bg="#1A7F3C", fg="white", font=("Arial", 11, "bold"),
                             relief="flat", padx=20, pady=8, cursor="hand2", state="disabled")
        self.btn.pack(pady=(10, 16))

    def _pick_pdf(self):
        path = filedialog.askopenfilename(
            title="Selecionar PDF",
            filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")]
        )
        if not path:
            return
        self.pdf_var.set(path)
        self.xlsx_var.set(path.rsplit(".", 1)[0] + ".xlsx")
        self.btn.config(state="normal")
        self.status_var.set("Pronto para converter.")

    def _pick_xlsx(self):
        path = filedialog.asksaveasfilename(
            title="Salvar Excel como",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if path:
            self.xlsx_var.set(path)

    def _start(self):
        pdf_in   = self.pdf_var.get()
        xlsx_out = self.xlsx_var.get()
        if not pdf_in or not xlsx_out:
            messagebox.showwarning("Atenção", "Selecione o PDF e o destino do Excel.")
            return
        self.btn.config(state="disabled")
        self.progress["value"] = 0
        self.status_var.set("Convertendo… aguarde.")
        threading.Thread(target=self._run, args=(pdf_in, xlsx_out), daemon=True).start()

    def _run(self, pdf_in, xlsx_out):
        try:
            def cb(current, total):
                pct = int(current / total * 100)
                self.after(0, lambda: self._set_progress(pct, current, total))
            parse_pdf(pdf_in, xlsx_out, progress_cb=cb)
            self.after(0, self._done, xlsx_out, None)
        except Exception as e:
            self.after(0, self._done, None, str(e))

    def _set_progress(self, pct, current, total):
        self.progress["value"] = pct
        self.status_var.set(f"Processando página {current} de {total}…")

    def _done(self, xlsx_out, error):
        self.progress["value"] = 100 if not error else 0
        self.btn.config(state="normal")
        if error:
            self.status_var.set(f"Erro: {error}")
            messagebox.showerror("Erro na conversão", error)
        else:
            self.status_var.set(f"✅  Concluído: {xlsx_out}")
            messagebox.showinfo("Concluído!", f"Excel salvo em:\n{xlsx_out}")

# ─── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()